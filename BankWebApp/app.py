from flask import Flask, render_template, request
import openpyxl
import os
from datetime import datetime
import re
import time
import tempfile
import hashlib

app = Flask(__name__)
FILE_NAME = "Data.xlsx"


# ---------------- EXCEL SETUP ----------------
def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "Name", "Address", "Email", "Phone",
            "Account No", "Balance", "Created At", "Last Updated",
            "PIN Hash", "PIN Salt"
        ])
        save_workbook_safe(wb, FILE_NAME)


# helper: save workbook with retries on PermissionError
def save_workbook_safe(wb, filename, retries=6, delay=0.5):
    """Try saving the workbook. If a PermissionError occurs (file open in Excel),
    attempt to save to a temporary file and replace the original. Returns True on success."""
    tmp_path = None
    for attempt in range(retries):
        try:
            wb.save(filename)
            return True
        except PermissionError:
            try:
                # create a temporary file in same directory
                fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
                os.close(fd)
                wb.save(tmp_path)
                os.replace(tmp_path, filename)
                return True
            except Exception:
                try:
                    if tmp_path and os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass
        except Exception:
            # for non-permission errors, do not retry
            break
        time.sleep(delay)
    return False


# ---------------- HOME PAGE ----------------
@app.route("/")
def index():
    return render_template("index.html", view='dashboard')


# ---------------- ADD CUSTOMER ----------------
@app.route("/add_customer", methods=["GET", "POST"])
def add_customer():
    if request.method == "POST":
        errors = {}
        name = request.form.get("name", "").strip()
        address = request.form.get("address", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        deposit_raw = request.form.get("deposit", "").strip()

        # Validate name (letters, spaces, hyphens, apostrophes)
        if not name or not re.match(r"^[A-Za-z\s'\-\.]{2,}$", name):
            errors['name'] = "Name must contain only letters, spaces or common name punctuation."

        # Validate email (simple check)
        if not email or not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            errors['email'] = "Enter a valid email address."

        # Validate phone (10 digits)
        if not phone.isdigit() or len(phone) != 10:
            errors['phone'] = "Phone must be a 10-digit number."

        # Validate deposit (allow values like '1000', '1,000', or '1000.00')
        deposit_raw_clean = deposit_raw.replace(',','')
        try:
            deposit_f = float(deposit_raw_clean)
            if deposit_f < 0:
                errors['deposit'] = "Deposit must be non-negative."
            elif not deposit_f.is_integer():
                errors['deposit'] = "Please enter a whole number (no decimals)."
            else:
                deposit = int(deposit_f)
        except Exception:
            errors['deposit'] = "Enter a valid deposit amount (numbers only)."

        if errors:
            return render_template("index.html", view='add_customer', errors=errors, form={'name': name, 'address': address, 'email': email, 'phone': phone, 'deposit': deposit_raw})

        # Validate PIN (must be 4 digits)
        pin = request.form.get('pin', '').strip()
        if not pin or not pin.isdigit() or len(pin) != 4:
            errors['pin'] = "Enter a 4-digit numeric PIN."

        if errors:
            # do not include PIN in the return form data for security
            return render_template("index.html", view='add_customer', errors=errors, form={'name': name, 'address': address, 'email': email, 'phone': phone, 'deposit': deposit_raw})

        # Load spreadsheet and check for duplicates (email, phone, or same name+address)
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            existing_acc = str(row[4].value).strip() if row[4].value is not None else ''
            existing_email = str(row[2].value).strip().lower() if row[2].value else ''
            existing_phone = str(row[3].value).strip() if row[3].value else ''
            existing_name = str(row[0].value).strip().lower() if row[0].value else ''
            existing_address = str(row[1].value).strip().lower() if row[1].value else ''

            if email and existing_email and email.lower() == existing_email:
                errors['email'] = f"Email already in use (Account: {existing_acc})"
            if phone and existing_phone and phone == existing_phone:
                errors['phone'] = f"Phone already in use (Account: {existing_acc})"
            if name and existing_name and address and existing_address and name.lower() == existing_name and address.lower() == existing_address:
                errors['duplicate'] = f"A customer with the same name and address already exists (Account: {existing_acc})."

        if errors:
            return render_template("index.html", view='add_customer', errors=errors, form={'name': name, 'address': address, 'email': email, 'phone': phone, 'deposit': deposit_raw})

        # Generate unique account number based on existing max
        base = 1000000000
        max_acc = 0
        for row in ws.iter_rows(min_row=2):
            try:
                val = int(row[4].value)
                if val > max_acc:
                    max_acc = val
            except Exception:
                continue

        if max_acc < base:
            accno = str(base)
        else:
            accno = str(max_acc + 1)

        # Hash the PIN with a random salt (store as hex)
        salt = os.urandom(8).hex()
        pin_hash = hashlib.pbkdf2_hmac('sha256', pin.encode(), salt.encode(), 100000).hex()
        time_now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        ws.append([
            name, address, email, phone,
            accno, deposit, time_now, time_now,
            pin_hash, salt
        ])

        if not save_workbook_safe(wb, FILE_NAME):
            errors = {'save': 'Could not save data file. Please close Data.xlsx if it is open and try again.'}
            return render_template("index.html", view='add_customer', errors=errors, form={'name': name, 'address': address, 'email': email, 'phone': phone, 'deposit': deposit})

        return render_template("index.html", view='add_customer', msg="Customer added successfully! The 4-digit PIN has been stored securely.", accno=accno, name=name, form={'deposit': deposit})

    return render_template("index.html", view='add_customer')


# ---------------- DEPOSIT MONEY ----------------
@app.route("/deposit", methods=["GET", "POST"])
def deposit():
    if request.method == "POST":
        accno = request.form["accno"].strip()
        amount = int(request.form["amount"])
        time_now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        pin = request.form.get('pin','').strip()
        for row in ws.iter_rows(min_row=2):
            cell_acc = row[4].value
            if cell_acc is not None and str(cell_acc).strip() == accno:
                # Verify PIN exists for account
                try:
                    stored_hash = str(row[8].value) if row[8].value is not None else ''
                    stored_salt = str(row[9].value) if row[9].value is not None else ''
                except IndexError:
                    stored_hash = ''
                    stored_salt = ''

                if not stored_hash or not stored_salt:
                    return render_template('index.html', view='deposit', msg='This account does not have a PIN set. Cannot proceed.', form={'accno': accno, 'amount': amount})

                # Validate provided PIN
                if not pin or not pin.isdigit() or len(pin) != 4:
                    return render_template('index.html', view='deposit', msg='Invalid PIN format. Enter 4 digits.', form={'accno': accno, 'amount': amount})

                check_hash = hashlib.pbkdf2_hmac('sha256', pin.encode(), stored_salt.encode(), 100000).hex()
                if check_hash != stored_hash:
                    return render_template('index.html', view='deposit', msg='Invalid PIN provided.', form={'accno': accno, 'amount': amount})

                # Safely parse existing balance and add amount
                old_balance = row[5].value
                try:
                    current = int(float(str(old_balance).replace(',',''))) if old_balance is not None else 0
                except Exception:
                    current = 0
                row[5].value = current + amount
                row[7].value = time_now
                if not save_workbook_safe(wb, FILE_NAME):
                    return render_template("index.html", view='deposit', msg="Could not save data file. Please close Data.xlsx if it is open and try again.", form={'accno': accno})
                return render_template(
                    "index.html",
                    view='deposit',
                    msg="Amount deposited successfully!",
                    balance=row[5].value
                )

        return render_template("index.html", view='deposit', msg="Account not found!")

    return render_template("index.html", view='deposit')


# ---------------- CHECK BALANCE ----------------
@app.route("/balance", methods=["GET", "POST"])
def balance():
    if request.method == "POST":
        accno = request.form["accno"].strip()
        pin = request.form.get('pin','').strip()

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            cell_acc = row[4].value
            if cell_acc is not None and str(cell_acc).strip() == accno:
                # retrieve stored PIN hash/salt
                try:
                    stored_hash = str(row[8].value) if row[8].value is not None else ''
                    stored_salt = str(row[9].value) if row[9].value is not None else ''
                except IndexError:
                    stored_hash = ''
                    stored_salt = ''

                if not stored_hash or not stored_salt:
                    return render_template('index.html', view='balance', msg='This account does not have a PIN set. Cannot show balance.', form={'accno': accno})

                if not pin or not pin.isdigit() or len(pin) != 4:
                    return render_template('index.html', view='balance', msg='Invalid PIN format. Enter 4 digits.', form={'accno': accno})

                check_hash = hashlib.pbkdf2_hmac('sha256', pin.encode(), stored_salt.encode(), 100000).hex()
                if check_hash != stored_hash:
                    return render_template('index.html', view='balance', msg='Invalid PIN provided.', form={'accno': accno})

                return render_template(
                    "index.html",
                    view='balance',
                    balance=row[5].value,
                    updated=row[7].value
                )

        return render_template("index.html", view='balance', msg="Account not found!")

    return render_template("index.html", view='balance')


# ---------------- VIEW CUSTOMER DETAILS ----------------
@app.route("/view_details", methods=["GET", "POST"])
def view_details():
    if request.method == "POST":
        accno = request.form["accno"].strip()
        pin = request.form.get('pin','').strip()

        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            cell_acc = row[4].value
            if cell_acc is not None and str(cell_acc).strip() == accno:
                # retrieve stored PIN hash/salt
                try:
                    stored_hash = str(row[8].value) if row[8].value is not None else ''
                    stored_salt = str(row[9].value) if row[9].value is not None else ''
                except IndexError:
                    stored_hash = ''
                    stored_salt = ''

                if not stored_hash or not stored_salt:
                    return render_template('index.html', view='view_details', msg='This account does not have a PIN set. Cannot show details.', form={'accno': accno})

                if not pin or not pin.isdigit() or len(pin) != 4:
                    return render_template('index.html', view='view_details', msg='Invalid PIN format. Enter 4 digits.', form={'accno': accno})

                check_hash = hashlib.pbkdf2_hmac('sha256', pin.encode(), stored_salt.encode(), 100000).hex()
                if check_hash != stored_hash:
                    return render_template('index.html', view='view_details', msg='Invalid PIN provided.', form={'accno': accno})

                return render_template("index.html", view='view_details', row=row)

        return render_template("index.html", view='view_details', msg="Account not found!")

    return render_template("index.html", view='view_details')


# ---------------- RUN APPLICATION ----------------
if __name__ == "__main__":
    init_excel()
    app.run(debug=True,port=1414)